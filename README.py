import os
import tkinter as tk
from tkinter import messagebox
from datetime import datetime

import pandas as pd
import win32com.client as win32
from openpyxl import Workbook, load_workbook


# ============================================================
# CONFIGURATION
# ============================================================

BASE_FOLDER = os.path.dirname(os.path.abspath(__file__))

HELPER_FILE = os.path.join(
    BASE_FOLDER,
    "Email_Bot_Helper.xlsx"
)

AUDIT_FILE = os.path.join(
    BASE_FOLDER,
    "Email_Bot_Audit.xlsx"
)


# ============================================================
# CHANGE THESE VALUES
# ============================================================

MAILBOX_NAME = "YOUR GROUP MAILBOX NAME"

GROUP_MAILBOX_EMAIL = "your-group-mailbox@jpmorgan.com"


# ============================================================
# OUTLOOK CATEGORIES
# ============================================================

DELAWARE_CATEGORY = "Delaware"

EMAIL_BOT_CATEGORY = "Email BOT"


# ============================================================
# INTERNAL JPM DOMAINS
# ============================================================

INTERNAL_DOMAINS = (
    "jpmorgan.com",
    "jpmchase.com",
    "jpmc.com",
)


# ============================================================
# ACKNOWLEDGEMENT MESSAGE
# ============================================================

ACKNOWLEDGEMENT_TEXT = (
    "Thank you for your request. "
    "We are looking into your request and will get back to you."
)


# ============================================================
# HELPER FILE
# ============================================================

def load_credit_contacts():

    if not os.path.exists(HELPER_FILE):

        raise FileNotFoundError(
            f"Helper file not found:\n\n{HELPER_FILE}"
        )

    df = pd.read_excel(HELPER_FILE)

    credit_columns = [
        col
        for col in df.columns
        if "credit" in str(col).lower()
    ]

    if not credit_columns:

        raise ValueError(
            "No Credit columns were found in the helper file."
        )

    credit_contacts = set()

    for col in credit_columns:

        for value in df[col].dropna():

            email = str(value).strip().lower()

            if "@" in email:
                credit_contacts.add(email)

    if not credit_contacts:

        raise ValueError(
            "No Credit email addresses were found in the helper file."
        )

    return credit_contacts


# ============================================================
# OUTLOOK CONNECTION
# ============================================================

def get_outlook_namespace():

    outlook = win32.Dispatch(
        "Outlook.Application"
    )

    namespace = outlook.GetNamespace(
        "MAPI"
    )

    return namespace


# ============================================================
# FIND GROUP MAILBOX INBOX
# ============================================================

def get_group_inbox(namespace):

    for mailbox in namespace.Folders:

        if (
            mailbox.Name.strip().lower()
            == MAILBOX_NAME.strip().lower()
        ):

            try:
                return mailbox.Folders["Inbox"]

            except Exception:

                raise ValueError(
                    f"Inbox not found under:\n{MAILBOX_NAME}"
                )

    raise ValueError(
        f"Mailbox not found in Outlook:\n\n{MAILBOX_NAME}"
    )


# ============================================================
# GET SENDER EMAIL
# ============================================================

def get_sender_email(mail):

    try:

        if mail.SenderEmailType == "EX":

            sender = mail.Sender

            if sender:

                exchange_user = sender.GetExchangeUser()

                if exchange_user:

                    smtp = exchange_user.PrimarySmtpAddress

                    if smtp:
                        return smtp.strip().lower()


                exchange_dl = (
                    sender.GetExchangeDistributionList()
                )

                if exchange_dl:

                    smtp = exchange_dl.PrimarySmtpAddress

                    if smtp:
                        return smtp.strip().lower()


        smtp = mail.SenderEmailAddress

        if smtp:
            return smtp.strip().lower()

    except Exception:
        pass

    return ""


# ============================================================
# GET RECIPIENT EMAIL
# ============================================================

def get_recipient_email(recipient):

    try:

        address_entry = recipient.AddressEntry

        if not address_entry:
            return ""


        if address_entry.Type == "EX":

            exchange_user = (
                address_entry.GetExchangeUser()
            )

            if exchange_user:

                smtp = exchange_user.PrimarySmtpAddress

                if smtp:
                    return smtp.strip().lower()


            exchange_dl = (
                address_entry.GetExchangeDistributionList()
            )

            if exchange_dl:

                smtp = exchange_dl.PrimarySmtpAddress

                if smtp:
                    return smtp.strip().lower()


        address = address_entry.Address

        if address:
            return address.strip().lower()

    except Exception:
        pass

    return ""


# ============================================================
# INTERNAL EMAIL VALIDATION
# ============================================================

def is_internal_email(email):

    if not email or "@" not in email:
        return False

    domain = email.rsplit("@", 1)[1].lower()

    return domain in INTERNAL_DOMAINS


# ============================================================
# CHECK ALL TO / CC RECIPIENTS
# ============================================================

def all_recipients_internal(mail):

    try:

        if mail.Recipients.Count == 0:
            return False

        for recipient in mail.Recipients:

            email = get_recipient_email(
                recipient
            )

            if not email:
                return False

            if not is_internal_email(email):
                return False

        return True

    except Exception:

        return False


# ============================================================
# CATEGORY FUNCTIONS
# ============================================================

def get_categories(mail):

    try:

        raw_categories = mail.Categories or ""

        return [
            category.strip()
            for category in raw_categories.split(",")
            if category.strip()
        ]

    except Exception:

        return []


def has_category(mail, category_name):

    categories = get_categories(mail)

    return any(
        category.lower() == category_name.lower()
        for category in categories
    )


# ============================================================
# ENSURE EMAIL BOT CATEGORY EXISTS
# ============================================================

def ensure_email_bot_category(namespace):

    try:

        categories = namespace.Categories

        for category in categories:

            if (
                category.Name.strip().lower()
                == EMAIL_BOT_CATEGORY.lower()
            ):
                return

        categories.Add(
            EMAIL_BOT_CATEGORY,
            0
        )

    except Exception as e:

        print(
            f"Could not create Email BOT category: {e}"
        )


# ============================================================
# APPLY EMAIL BOT CATEGORY
# ============================================================

def apply_email_bot_category(mail):

    categories = get_categories(mail)

    if not any(
        category.lower() == EMAIL_BOT_CATEGORY.lower()
        for category in categories
    ):

        categories.append(
            EMAIL_BOT_CATEGORY
        )

    mail.Categories = ", ".join(
        categories
    )

    mail.Save()


# ============================================================
# CREATE AUDIT FILE
# ============================================================

def create_audit_file():

    if os.path.exists(AUDIT_FILE):
        return

    wb = Workbook()

    ws = wb.active

    ws.title = "Acknowledgement Audit"

    ws.append([
        "Received Time",
        "Sender Name",
        "Sender Email",
        "Subject",
        "Message ID",
        "Action",
        "Acknowledgement Sent Time",
        "Run Date Time"
    ])

    wb.save(
        AUDIT_FILE
    )


# ============================================================
# DUPLICATE CHECK FROM AUDIT
# ============================================================

def already_in_audit(message_id):

    create_audit_file()

    wb = load_workbook(
        AUDIT_FILE,
        read_only=True
    )

    ws = wb[
        "Acknowledgement Audit"
    ]

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        existing_id = row[4]

        if existing_id:

            if str(existing_id) == str(message_id):

                wb.close()

                return True

    wb.close()

    return False


# ============================================================
# APPEND TO AUDIT
# ============================================================

def append_to_audit(
    mail,
    sender_email
):

    create_audit_file()

    wb = load_workbook(
        AUDIT_FILE
    )

    ws = wb[
        "Acknowledgement Audit"
    ]

    now = datetime.now()

    ws.append([
        mail.ReceivedTime,
        mail.SenderName,
        sender_email,
        mail.Subject,
        mail.EntryID,
        "Acknowledgement Sent",
        now,
        now
    ])

    wb.save(
        AUDIT_FILE
    )


# ============================================================
# CREATE ACKNOWLEDGEMENT
# ============================================================

def create_acknowledgement(
    mail,
    run_mode
):

    reply = mail.Reply()

    # CC group mailbox
    reply.CC = GROUP_MAILBOX_EMAIL

    # Preserve original email thread
    original_body = reply.HTMLBody

    acknowledgement_html = f"""
    <p>{ACKNOWLEDGEMENT_TEXT}</p>
    <br>
    """

    reply.HTMLBody = (
        acknowledgement_html
        + original_body
    )

    if run_mode == "Draft":

        reply.Save()

        return "Draft"

    else:

        reply.Send()

        return "Sent"


# ============================================================
# PROCESS PHASE 1 EMAILS
# ============================================================

def process_emails(run_mode):

    credit_contacts = (
        load_credit_contacts()
    )

    namespace = (
        get_outlook_namespace()
    )

    ensure_email_bot_category(
        namespace
    )

    inbox = get_group_inbox(
        namespace
    )

    items = inbox.Items

    items.Sort(
        "[ReceivedTime]",
        True
    )

    emails_found = 0

    emails_sent = 0

    today = datetime.now().date()


    # ========================================================
    # LOOP THROUGH INBOX
    # ========================================================

    for mail in items:

        try:

            # Outlook MailItem
            if mail.Class != 43:
                continue


            # =================================================
            # CHECK 1 - TODAY'S EMAIL ONLY
            # =================================================

            try:

                received_date = (
                    mail.ReceivedTime.date()
                )

            except Exception:

                continue


            # Since newest emails are first,
            # stop once yesterday is reached
            if received_date < today:
                break


            if received_date != today:
                continue


            # =================================================
            # CHECK 2 - MUST HAVE DELAWARE CATEGORY
            # =================================================

            if not has_category(
                mail,
                DELAWARE_CATEGORY
            ):
                continue


            # =================================================
            # CHECK 3 - ALREADY PROCESSED?
            # =================================================

            if has_category(
                mail,
                EMAIL_BOT_CATEGORY
            ):
                continue


            # =================================================
            # CHECK 4 - GET SENDER
            # =================================================

            sender_email = (
                get_sender_email(mail)
            )

            if not sender_email:
                continue


            # =================================================
            # CHECK 5 - SENDER MUST BE IN HELPER
            # =================================================

            if sender_email not in credit_contacts:
                continue


            # =================================================
            # CHECK 6 - SENDER MUST BE INTERNAL
            # =================================================

            if not is_internal_email(
                sender_email
            ):
                continue


            # =================================================
            # CHECK 7 - ALL TO / CC MUST BE INTERNAL
            # =================================================

            if not all_recipients_internal(
                mail
            ):
                continue


            # =================================================
            # CHECK 8 - AUDIT DUPLICATE CONTROL
            # =================================================

            if already_in_audit(
                mail.EntryID
            ):
                continue


            # =================================================
            # EMAIL IS ELIGIBLE
            # =================================================

            emails_found += 1


            # =================================================
            # CREATE / SEND RESPONSE
            # =================================================

            result = create_acknowledgement(
                mail,
                run_mode
            )


            # =================================================
            # ONLY AFTER AUTO-SEND
            # =================================================

            if result == "Sent":

                apply_email_bot_category(
                    mail
                )

                append_to_audit(
                    mail,
                    sender_email
                )

                emails_sent += 1


        except Exception as e:

            print(
                "Error processing email:",
                getattr(
                    mail,
                    "Subject",
                    ""
                ),
                e
            )


    return (
        emails_found,
        emails_sent
    )


# ============================================================
# GUI
# ============================================================

root = tk.Tk()

root.title(
    "Email Bot Automation"
)

root.geometry(
    "520x340"
)

root.resizable(
    False,
    False
)


# ============================================================
# TITLE
# ============================================================

title_label = tk.Label(
    root,
    text="EMAIL BOT AUTOMATION",
    font=(
        "Segoe UI",
        19,
        "bold"
    )
)

title_label.pack(
    pady=(
        30,
        25
    )
)


# ============================================================
# RUN MODE
# ============================================================

mode_frame = tk.Frame(
    root
)

mode_frame.pack(
    pady=10
)


tk.Label(
    mode_frame,
    text="Run Mode:",
    font=(
        "Segoe UI",
        11,
        "bold"
    )
).pack(
    side="left",
    padx=10
)


run_mode_var = tk.StringVar(
    value="Draft"
)


tk.Radiobutton(
    mode_frame,
    text="Draft",
    variable=run_mode_var,
    value="Draft",
    font=(
        "Segoe UI",
        10
    )
).pack(
    side="left",
    padx=10
)


tk.Radiobutton(
    mode_frame,
    text="Auto-Send",
    variable=run_mode_var,
    value="Auto-Send",
    font=(
        "Segoe UI",
        10
    )
).pack(
    side="left",
    padx=10
)


# ============================================================
# SUMMARY VARIABLES
# ============================================================

emails_found_var = tk.StringVar(
    value="0"
)

emails_sent_var = tk.StringVar(
    value="0"
)


# ============================================================
# RUN FUNCTION
# ============================================================

def run_email_bot():

    run_mode = run_mode_var.get()

    emails_found_var.set("0")

    emails_sent_var.set("0")

    root.update_idletasks()


    if run_mode == "Auto-Send":

        confirm = messagebox.askyesno(
            "Confirm Auto-Send",
            "Are you sure you want to send "
            "the acknowledgement emails?"
        )

        if not confirm:
            return


    try:

        (
            emails_found,
            emails_sent
        ) = process_emails(
            run_mode
        )


        emails_found_var.set(
            str(emails_found)
        )

        emails_sent_var.set(
            str(emails_sent)
        )


        if run_mode == "Draft":

            messagebox.showinfo(
                "Draft Run Completed",
                f"{emails_found} eligible email(s) found.\n\n"
                "Acknowledgement drafts have been created."
            )

        else:

            messagebox.showinfo(
                "Email Bot Completed",
                f"{emails_sent} acknowledgement email(s) "
                "sent successfully."
            )


    except Exception as e:

        messagebox.showerror(
            "Email Bot Error",
            str(e)
        )


# ============================================================
# RUN BUTTON
# ============================================================

run_button = tk.Button(
    root,
    text="RUN EMAIL BOT",
    command=run_email_bot,
    font=(
        "Segoe UI",
        11,
        "bold"
    ),
    width=22,
    height=2
)

run_button.pack(
    pady=15
)


# ============================================================
# SUMMARY
# ============================================================

summary_frame = tk.LabelFrame(
    root,
    text=" Run Summary ",
    font=(
        "Segoe UI",
        10,
        "bold"
    ),
    padx=30,
    pady=12
)

summary_frame.pack(
    fill="x",
    padx=80,
    pady=10
)


tk.Label(
    summary_frame,
    text="Emails Found:",
    font=(
        "Segoe UI",
        11
    )
).grid(
    row=0,
    column=0,
    sticky="w",
    pady=4
)


tk.Label(
    summary_frame,
    textvariable=emails_found_var,
    font=(
        "Segoe UI",
        11,
        "bold"
    )
).grid(
    row=0,
    column=1,
    padx=70
)


tk.Label(
    summary_frame,
    text="Emails Sent:",
    font=(
        "Segoe UI",
        11
    )
).grid(
    row=1,
    column=0,
    sticky="w",
    pady=4
)


tk.Label(
    summary_frame,
    textvariable=emails_sent_var,
    font=(
        "Segoe UI",
        11,
        "bold"
    )
).grid(
    row=1,
    column=1,
    padx=70
)


# ============================================================
# START GUI
# ============================================================

root.mainloop()
